#!/usr/bin/env python3
# coding: utf-8
import rclpy
from rclpy.node import Node
from rclpy.clock import Clock
from rclpy.action import ActionClient
# from std_msgs.msg import String
# from builtin_interfaces.msg import Duration
from control_msgs.action import FollowJointTrajectory
from builtin_interfaces.msg import Duration, Time
from trajectory_msgs.msg import JointTrajectory , JointTrajectoryPoint
from sensor_msgs.msg import JointState
from control_msgs.msg import JointTrajectoryControllerState
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
# import pandas as pd
from openpyxl import load_workbook
curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
# from my_package.
from robot_communicate.msg import JointTorque
from roboticstoolbox.robot.Robot import Robot

class modular_robot_6dof(DHRobot):
    def __init__(self):
        self.length_1 = 0.1
        self.length_2 = 0.1
        self.length_3 = 0.1
        symbolic = False
        if symbolic:
            import spatialmath.base.symbolic as sym
            zero = sym.zero()
            pi = sym.pi()
        else:
            from math import pi
            zero = 0.0

        deg = pi / 180
        inch = 0.0254
        a = [0, 0.26754, 0.23246, 0, 0, 0] 
        d = [0.176, 0, 0, 0.0806, 0.0806, 0.0549999394444907]
        alpha = [pi/2, 0, zero, pi/2, -pi/2, zero]
        # mass = [0.93428, 1.4067641114285714, 1.1306573171428573, 0.36924, 0.36821, 0.0259770442543706]
        mass = [0.93428, 1.07602, 1.00039, 0.36924, 0.36821, 0.0259770442543706]
        G= [-80,-80,-80,-50,-50,-50]   # gear ratio
        G= [-1,-1,-1,-1,-1,-1]   # gear ratio
        B = 10.0
        center_of_mass = [[3.305e-06, 0.0036146, 0.029085], [0.2128581564957513, 1.646135446113019e-06, 0.012407278181240242], [0.18456174031042727, 1.661355600613535e-07, -0.010833273129287313], [2.2204e-16, 0.01039, 0.02404], [0.028614, 0.010457, -3.3307e-16], [-4.44089209850063e-16, 5.37938091366846e-07, -0.00700006055550936]]
        # signed from a 3 3 matrix or a 6-element vector interpretted as Ixx Iyy Izz Ixy Iyz Ixz
        inertia = [[0.00048112, 0.00018986, 0.0004948, -7.0827e-08, 4.1821e-06, -3.5566e-10], [0.0009354044506587581, 0.0034611999079844536, 0.002717487568255342, 3.9417575092874916e-08, -1.4069358691068503e-08, -0.0006551621120647835], [0.0007066546120021806, 0.002750976174570803, 0.0022186586319623814, 1.2966478903714108e-08, -5.656649319035686e-10, 0.0004032273103514373], [9.4187e-05, 4.476e-05, 0.0001027, 1.2869e-19, 5.3346e-07, -1.4501e-16], [0.00010133, 4.3878e-05, 9.3658e-05, 5.3346e-07, -8.1817e-20, 1.4501e-16], [2.20534281951167e-06, 2.20534281951168e-06, 3.9777349014505e-06, 1.78671012311437e-22, 2.32583060538047e-20, -8.79317315621994e-22]]
        # rad = deg/180*pi
        q_lim = [
            [-pi,pi],
            [0,pi],
            [-150/180*pi,150/180*pi],
            [-pi,pi],
            [-pi,pi],
            [-pi,pi]
        ]
        links = []

        for j in range(6):
            link = RevoluteDH(
                d=d[j],
                a=a[j],
                alpha=alpha[j],
                m=mass[j],
                r=center_of_mass[j],
                I=inertia[j],
                G=G[j],
                # B=B[j]
                qlim=q_lim[j]
            )
            links.append(link)
    
        super().__init__(
            links,
            name="robot",
            manufacturer="Robotics",
            keywords=('dynamics', 'symbolic'),
            symbolic=symbolic
        )
        # zero angles
        self.addconfiguration("qz", np.array([0, 0, 0, 0, 0, 0]))
        # horizontal along the x-axis
        self.addconfiguration("qr", np.r_[180, 0, 0, 0, 90, 0]*deg)

        # nominal table top picking pose
        self.addconfiguration("qn", np.array([0, 0, 0, 0, 0, 0]))
class modular_robot_6dof_ets(Robot):
    def __init__(self):

        links, name, urdf_string, urdf_filepath = self.URDF_read(
            "./src/single_arm_6dof_description/urdf/single_arm_6dof.urdf.xacro"
        )
        # for link in links:
        #     print(link)

        super().__init__(
            links,
            name=name.upper(),
            manufacturer="Universal Robotics",
            # gripper_links=links[7],
            urdf_string=urdf_string,
            urdf_filepath=urdf_filepath,
        )

        self.qr = np.array([np.pi, 0, 0, 0, np.pi / 2, 0])
        self.qz = np.zeros(6)

        self.addconfiguration("qr", self.qr)
        self.addconfiguration("qz", self.qz)

        # sol=robot.ikine_LM(SE3(0.5, -0.2, 0.2)@SE3.OA([1,0,0],[0,0,-1]))
        self.addconfiguration_attr(
            "qn",
            np.array(
                [
                    -7.052413e-01,
                    3.604328e-01,
                    -1.494176e00,
                    1.133744e00,
                    -7.052413e-01,
                    0,
                ]
            ),
        )
        self.addconfiguration_attr("q1", [0, -np.pi / 2, np.pi / 2, 0, np.pi / 2, 0])

class RobotSubscriber(Node):

    def __init__(self, robot):
        super().__init__('robot_subscriber')
        self.subscription = self.create_subscription(
            JointState,
            '/joint_states',
            self.listener_callback,
            10)
        
        self.robot = robot
        # 在初始化函数中创建发布者
        self.torque_publisher = self.create_publisher(JointTorque, 'joint_torque_topic', 10)

        # 在适当的地方获取关节扭矩信息，并发布消息
        self.joint_torque_msg = JointTorque()
        self.joint_torque_msg.sim_joint_names = ['joint1', 'joint2', 'joint3', 'joint4', 'joint5', 'joint6']
        self.joint_torque_msg.torques = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        self.joint_torque_msg.real_current = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        self.joint_torque_msg.real_joint_names = ['joint1', 'joint2', 'joint3', 'joint4', 'joint5', 'joint6']
        self.joint_torque_msg.pub_position = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        # 发布关节扭矩消息
        self.torque_publisher.publish(self.joint_torque_msg)
        # 上一次关节状态的变量
        self.previous_position = None
        self.previous_velocity = None
        self.subscription  # prevent unused variable warning

    def listener_callback(self, msg):
        # TODO: fixed sub joint name 順序
        # 按照特定的顺序调整关节数据
        # 获取关节名称列表
        joint_names = msg.name
        desired_order = ['joint1', 'joint2', 'joint3', 'joint4', 'joint5', 'joint6']
        ordered_positions = []
        ordered_velocities = []
        ordered_effort = []
        for joint_name in desired_order:
            if joint_name in joint_names:
                index = joint_names.index(joint_name)
                ordered_positions.append(msg.position[index])
                ordered_velocities.append(msg.velocity[index])
                ordered_effort.append(msg.effort[index])
            else:
                self.get_logger().info("Joint {} not found in joint_state message.".format(joint_name))
        # 获取当前的关节状态
        current_name = desired_order
        current_position = ordered_positions
        current_velocity = ordered_velocities
        current_effort = ordered_effort
        current_time = self.get_clock().now()
        # print("current_time:", current_time)
        # TODO: 計算動力學扭矩
            # 位置切換回
        # current_position = [-msg.position[0], -(msg.position[1]+1.57), msg.position[2], -(msg.position[3]+1.57), -msg.position[4], msg.position[5]]
        # current_velocity = [-msg.velocity[0], -(msg.velocity[1]), msg.velocity[2], -(msg.velocity[3]), -msg.velocity[4], msg.velocity[5]]
            # 速度 正負
            # 加速度正負
        
        current_position = [current_position[0], current_position[1], -current_position[2], current_position[3], current_position[4], current_position[5]]
        current_velocity = [current_velocity[0], current_velocity[1], -current_velocity[2], current_velocity[3], current_velocity[4], current_velocity[5]]
        # current_effort = [msg.real_current[0], msg.real_current[1], msg.real_current[2], msg.real_current[3], msg.real_current[4], msg.real_current[5]]
        # 计算加速度
        if self.previous_position is not None and self.previous_velocity is not None:
            # 获取时间步长
            time_step = (current_time - self.previous_time).nanoseconds / 1e9
            # print("time_step:", time_step)
            # 计算位置变化率
            position_change = [
                (current_position[i] - self.previous_position[i]) / time_step
                for i in range(len(current_position))
            ]
            # 计算速度变化率
            velocity_change = [
                (current_velocity[i] - self.previous_velocity[i]) / time_step
                for i in range(len(current_velocity))
            ]
            # 计算加速度
            acceleration = [
                (velocity_change[i] - self.previous_velocity[i]) / time_step
                for i in range(len(velocity_change))
            ]
            # torque = self.robot.rne(msg.position.tolist(),msg.velocity.tolist(),acceleration)
            # torque = self.robot.rne(current_position,current_velocity,[0,0,0,0,0,0])
            # 斜率: 6.25
            # 截距: -6.375
            # 斜率: 6.521739130434783
            # 截距: -8.130434782608699
            # 斜率: 4.499999999999998
            # 截距: -1.4999999999999964
            # for i in range(len(current_effort)):
            #     if current_effort[i] < 0.0:
            #         if i == 0 or i == 1:
            #             current_effort[i] = -6.25*current_effort[i] + 6.375
            #         elif i == 2 or i == 3:
            #             current_effort[i] = -6.521739130434783*current_effort[i] + 8.130434782608699
            #         elif i == 4 or i == 5:
            #             current_effort[i] = -4.499999999999998*current_effort[i] + 1.4999999999999964
            #     elif current_effort[i] > 0.0:
            #         if i == 0 or i == 1:
            #             current_effort[i] = 6.25*current_effort[i] - 6.375
            #         elif i == 2 or i == 3:
            #             current_effort[i] = 6.521739130434783*current_effort[i] -8.130434782608699
            #         elif i == 4 or i == 5:
            #             current_effort[i] = 4.499999999999998*current_effort[i] -1.4999999999999964
                        
            torque = self.robot.rne(current_position,current_velocity,acceleration, gravity=[0, 0, -9.81])
            torque_list = [torque[0],torque[1],-torque[2],torque[3],torque[4],torque[5]]
            
            # self.joint_torque_msg.torques = torque.to_list()
            self.joint_torque_msg.torques = torque_list
            self.joint_torque_msg.real_current = [current_effort[0]*5/4,current_effort[1]*5/4,current_effort[2]*5/4,current_effort[3],current_effort[4], current_effort[5]]
            self.joint_torque_msg.real_joint_names = [current_name[0],current_name[1],current_name[2],current_name[3],current_name[4], current_name[5]]
            self.joint_torque_msg.pub_position = [current_position[0],current_position[1],current_position[2],current_position[3],current_position[4], current_position[5]]
            # 发布关节扭矩消息
            self.torque_publisher.publish(self.joint_torque_msg)
        # 更新上一次的状态
        self.previous_position = current_position
        self.previous_velocity = current_velocity
        self.previous_time = current_time
        
        
        
def main(args=None):
    rclpy.init(args=args)
    robot = modular_robot_6dof()
    payload = 1.1
    payload_position = [0, 0, 0.04]
    robot.payload(payload, payload_position) # set payload
    # robot_ets = modular_robot_6dof_ets()
    robot_subscriber = RobotSubscriber(robot)
    
    # robot.teach()

    rclpy.spin(robot_subscriber)
    robot_subscriber.destroy_node()
    rclpy.shutdown()
    
if __name__ == '__main__':    # pragma nocover
    main()
    
    # df = load_workbook("./xlsx/task_point_6dof_tested_ori_random.xlsx")
    # sheets = df.worksheets
    # sheet1 = sheets[0]
    # rows = sheet1.rows
    # cols = sheet1.columns
    # T_tmp = []
    # T_traj = []
    # ik_q_traj = []
    # score = []
    # ratio_over = 0
    # torque_over = 0
    # num_torque = np.array([np.zeros(shape=6)])
    # total_time = 20
    # # 采样间隔
    # sample_interval = 0.2
    # manipulability_index = []
    # i = 0
    # false_done = False
    # count = 0
    # max_diff = []
    # max_diff_tol = 0
    # traj_time = []
    # diff = np.array([np.zeros(shape=6)])
    # for row in rows:
    #     row_val = [col.value for col in row]
    #     T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
    #     ik_q = robot.ikine_LMS(T=T_tmp[i])

    #     if ik_q.success == True:
    #         count += 1
    #         T_traj.append(T_tmp[i])
    #         ik_q_traj.append(ik_q.q)
    #         manipulability_index.append(robot.manipulability(q=ik_q.q))
    #         print("ik_q.q",ik_q.q)
    #         if count >= 2: # 兩個點位以上開始計算
    #             diff = diff + np.abs(np.subtract(ik_q_traj[-2], ik_q_traj[-1]))
    #             max_diff_tol = max_diff_tol + np.max(diff)
    #             max_diff.append(np.max(diff))
    #             print(max_diff)
    #     i = i + 1
    # for k in range(len(max_diff)):
    #     traj_time.append(max_diff[k] / max_diff_tol * total_time)
    #     print("traj_time",traj_time)
    # for m in range(len(max_diff)):
    #     time_vector = np.linspace(0, traj_time[m], int(traj_time[m]/sample_interval) + 1)
    #     traj = robot.jtraj(T_traj[m],T_traj[m+1],time_vector)
    #     print(traj.s)
    #     if np.amax(traj.sd) > 3.04:
    #         ratio_over = ratio_over + 1
    #     torque = robot.rne(traj.s,traj.sd,traj.sdd)
    #     row = abs(torque[:,1]) # 取出第2行
    #     result_2 = row[row > 44.7] # 取出大于阈值的数字
    #     row = abs(torque[:,2]) # 取出第3行
    #     result_3 = row[row > 44.7] # 取出大于阈值的数字
    #     if len(result_2)>0 or len(result_3) >0:
    #         torque_over = torque_over + 1
    #     num_torque = np.append(num_torque, torque)
    # total_energy = 0
    # for j in range(len(num_torque)):
    #     energy = abs(num_torque[j]) * sample_interval
    #     total_energy += energy
            
        # if count == 0:
        #     return(0, 0, 0,0,0)
        # else:
        #     final_score = count / i
            # if count == 1:
            #     return(0, 0, 0, final_score, manipulability_index[0]) # 回傳 manipulability[0]
            # else:
            #     return(ratio_over, torque_over, total_energy, final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均


    # robot.plot(q=q_test, backend='pyplot', dt = 10)
    
    # robot.teach()
    # manipulability teach view
    # robot.teach(limits= [-0.5, 0.5, -0.5, 0.5, -0, 1],vellipse=True)
    
    '''
    # manipulability test point
    q0 =  np.r_[0,90,0,23,90,0]*deg
    q0 =  np.r_[0,76,-60,66,90,0]*deg
    vellipse = robot.vellipse(q=q0)
    robot.plot_ellipse(ellipse = vellipse)
    print(robot.manipulability(q=q0))
    '''

    # print(robot.fkine_path(q) * sm.SE3(0, 0, 0.04))

    # T = robot.fkine(q)
    # t = np.round(T.t, 3)
    # r = np.round(T.rpy(), 3)
    '''# print(robot.q)
 
    # robot.ikine_6s
    # robot.ikine_global
    # robot.ikine_LMS
    # robot.ikine_mmc
    T = SE3(0.7, 0.2, 0.1) * SE3.RPY([0, 1, 0])
    print(T)
    print(robot.ikine_LM(T=T))
    print(robot.ikine_LMS(T=T))

    print(robot.ikine_LM(T=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    print(robot.ikine_LMS(T=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    
    print(robot.manipulability(q=q))
    # print(robot.manipulability(J=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    
'''

'''
    T_tmp = []
    T_tmp.append(SE3(0.25, 0.113, 0.199) * SE3.RPY([np.deg2rad(-173), np.deg2rad(-59), np.deg2rad(-147)]))
    T_tmp.append(SE3(-0.06, -0.09, 0.20) * SE3.RPY([np.deg2rad(-79), np.deg2rad(27), np.deg2rad(-99)]))

    print(robot.ikine_LM(T=T_tmp[0]))
    print(robot.ikine_LMS(T=T_tmp[1]))
'''

'''
    robot.teach(limits= [-1, 1, -1, 1, -1, 1],vellipse=True)

    
    # import xlsx
    df = load_workbook("./xlsx/task_point_6dof.xlsx")
    sheets = df.worksheets
    sheet1 = sheets[0]
    rows = sheet1.rows
    cols = sheet1.columns
    robot.payload( 1.5, [0,0,0.04])  # set payload
    
    T_tmp = []
    num_torque = np.array([np.zeros(shape=6)])
    manipulability_index = []
    i = 0
    for row in rows:
        row_val = [col.value for col in row]
        T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
        # print(T_tmp[i])
        if i >= 1:
            # 完成动作的总时间
            total_time = 20
            # 采样间隔
            sample_interval = 0.056
            # 生成时间向量
            traj_time = total_time/10
            time_vector = np.linspace(0, traj_time, int(traj_time/sample_interval) + 1)
            # t=[0:0.056:2]
            traj = robot.jtraj(T_tmp[i-1],T_tmp[i],time_vector)
            # print(traj.s)
            # print(traj.sd)
            # # print(traj.sdd)
            # print(traj.t)
            # robot.plot(traj.s)
            # np.append(torque, load, axis=0)
            # print(traj.sd)
            # print(traj.sd[0])
            if np.amax(traj.sd) > 3.04:
                print("Fail")
            torque = robot.rne(traj.s,traj.sd,traj.sdd)
            if np.amax(torque) > 44.7:
                print(torque)
            num_torque = np.append(num_torque, torque)
            
        ik_q = robot.ikine_LMS(T=T_tmp[i])
        # print(ik_q)
        ik_np = np.array(ik_q.q)
        # robot.plot(q=ik_np, backend='pyplot', dt = 1)
        # if ik_q.success == True:
        #     manipulability_index.append(robot.manipulability(q=ik_q.q))
            # robot.plot_ellipse()
            # ik_np = np.array(ik_q.q)
            # print(ik_np)
            # robot.plot(q=ik_np, backend='pyplot', dt = 1)
        i = i + 1
    # print(np.mean(manipulability_index)) # manipulability 取平均
    time_interval = 0.056 # in seconds
    time_total = 20 # in seconds
    num_samples = int(time_total / time_interval)

    # Example motor power data
    motor_power = num_torque

    total_energy = 0
    for i in range(num_samples):
        energy = motor_power[i] * time_interval
        total_energy += energy
    print(total_energy)

    import numpy as np

    # num_actions = 12
    # one_hot_array = np.eye(num_actions)
    # print(one_hot_array)
    # action = 5
    # one_hot_vector = one_hot_array[action, :]
    # action = 8
    # one_hot_vector = one_hot_array[action, :]
    # print(one_hot_vector)

    matrix = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
    # threshold = 5
    # result = np.where(matrix > threshold, axis=0)
    # print(result)
    # row_max = np.amax(matrix, axis=1)
    # print(row_max)

    # if np.amax(matrix, axis=1) > 5:
    #     print("false")
    m = 3 # 行数，从0开始算
    threshold = 9 # 阈值

    row = matrix[:,m-1] # 取出第m-1行
    result = row[row > threshold] # 取出大于阈值的数字
    print(len(result))
'''